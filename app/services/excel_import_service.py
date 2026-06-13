from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.services.template_service import HEADERS
from app.utils.validators import ImportedQuestion, OPTION_KEYS, validate_question


@dataclass
class ImportResult:
    file_name: str
    total_count: int
    valid_count: int
    error_count: int
    duplicate_count: int
    questions: list[ImportedQuestion]
    errors: list[str]


class ExcelImportService:
    def import_file(self, file_path: str | Path) -> ImportResult:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在：{path}")
        if path.suffix.lower() != ".xlsx":
            raise ValueError("只支持 .xlsx 格式的 Excel 文件")

        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except PermissionError as exc:
            raise PermissionError("Excel 文件可能正在被占用，请关闭后重试") from exc
        except (InvalidFileException, OSError, Exception) as exc:
            raise ValueError(f"Excel 文件无法读取，可能格式损坏：{exc}") from exc

        if "题库数据" not in wb.sheetnames:
            raise ValueError("缺少工作表“题库数据”，请使用标准模板")
        ws = wb["题库数据"]
        header = [self._cell_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        missing = [item for item in HEADERS if item not in header]
        if missing:
            raise ValueError(f"表头缺失：{', '.join(missing)}")
        indexes = {name: header.index(name) for name in HEADERS}

        questions: list[ImportedQuestion] = []
        errors: list[str] = []
        seen_texts: set[str] = set()
        duplicate_count = 0
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(self._cell_text(v) == "" for v in row):
                continue
            item = ImportedQuestion(
                row_number=row_number,
                question_type=self._value(row, indexes["题型"]),
                question_text=self._value(row, indexes["题目"]),
                options={key: self._value(row, indexes[f"{key}选项"]) for key in OPTION_KEYS},
                correct_answer=self._value(row, indexes["正确答案"]),
                analysis=self._value(row, indexes["解析"]),
                chapter=self._value(row, indexes["章节"]),
                difficulty=self._value(row, indexes["难度"]),
            )
            validate_question(item, seen_texts)
            if any("重复题目" in err for err in item.errors):
                duplicate_count += 1
            for err in item.errors:
                errors.append(f"第 {row_number} 行：{err}")
            questions.append(item)

        if not questions:
            raise ValueError("Excel 中没有可读取的题目")
        return ImportResult(
            file_name=path.name,
            total_count=len(questions),
            valid_count=sum(1 for q in questions if q.is_valid),
            error_count=sum(1 for q in questions if not q.is_valid),
            duplicate_count=duplicate_count,
            questions=questions,
            errors=errors,
        )

    def _value(self, row: tuple, index: int) -> str:
        return self._cell_text(row[index] if index < len(row) else "")

    def _cell_text(self, value: object) -> str:
        return "" if value is None else str(value).strip()

