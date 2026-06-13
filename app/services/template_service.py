from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import TEMPLATES_DIR
from app.utils.file_utils import copy_file

TEMPLATE_NAME = "刷题软件题库导入模板.xlsx"
SAMPLE_NAME = "刷题软件示例题库.xlsx"
HEADERS = ["题型", "题目", "A选项", "B选项", "C选项", "D选项", "E选项", "F选项", "正确答案", "解析", "章节", "难度"]


class TemplateService:
    def __init__(self) -> None:
        TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

    def ensure_templates(self) -> None:
        blank = TEMPLATES_DIR / TEMPLATE_NAME
        sample = TEMPLATES_DIR / SAMPLE_NAME
        if not blank.exists():
            self._create_workbook(blank, include_samples=False)
        if not sample.exists():
            self._create_workbook(sample, include_samples=True)

    def copy_template(self, sample: bool, destination: str | Path) -> None:
        self.ensure_templates()
        src = TEMPLATES_DIR / (SAMPLE_NAME if sample else TEMPLATE_NAME)
        copy_file(src, destination)

    def _create_workbook(self, path: Path, include_samples: bool) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "题库数据"
        ws.append(HEADERS)
        if include_samples:
            rows = [
                ["单选题", "HTML 中定义段落的标签是？", "<p>", "<br>", "<div>", "<span>", "", "", "A", "p 标签用于定义段落", "HTML基础", "简单"],
                ["多选题", "以下哪些属于 HTML 标签？", "<p>", "<div>", "color", "<h1>", "", "", "ABD", "p、div、h1 都是 HTML 标签", "HTML基础", "中等"],
                ["判断题", "CSS 可以设置网页样式。", "", "", "", "", "", "", "正确", "CSS 用于控制网页样式", "CSS基础", "简单"],
                ["填空题", "HTML 文件的扩展名通常是____。", "", "", "", "", "", "", "html|htm", "两个答案都可接受", "HTML基础", "简单"],
            ]
            for row in rows:
                ws.append(row)

        self._style_data_sheet(ws)
        help_ws = wb.create_sheet("填写说明")
        help_rows = [
            ["题型支持", "单选题、单选；多选题、多选；判断题、判断；填空题、填空"],
            ["单选题答案", "填写 A"],
            ["多选题答案", "填写 ABD 或 A,B,D"],
            ["判断题答案", "正确、错误、对、错、√、×、True、False、1、0"],
            ["填空题答案", "多个可接受答案用竖线分隔，例如 html|htm"],
            ["难度", "简单、中等、困难；解析、章节、难度允许为空"],
            ["导入提示", "请保留工作表名称“题库数据”和第一行表头，不要合并单元格"],
        ]
        for row in help_rows:
            help_ws.append(row)
        help_ws.column_dimensions["A"].width = 18
        help_ws.column_dimensions["B"].width = 80
        for row in help_ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        wb.save(path)

    def _style_data_sheet(self, ws) -> None:
        widths = [12, 36, 16, 16, 16, 16, 16, 16, 16, 32, 16, 12]
        header_fill = PatternFill("solid", fgColor="DDEEFF")
        answer_fill = PatternFill("solid", fgColor="FFF2CC")
        thin = Side(style="thin", color="C9D6E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:L{max(ws.max_row, 2)}"
        for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 50), min_col=1, max_col=len(HEADERS)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center" if cell.row == 1 else "left", vertical="top", wrap_text=True)
                cell.border = border
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = answer_fill if cell.column == 9 else header_fill

