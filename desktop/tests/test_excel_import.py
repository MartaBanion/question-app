from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.excel_import_service import ExcelImportService
from app.services.template_service import HEADERS


def make_xlsx(path: Path, rows: list[list[str]]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "题库数据"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    wb.create_sheet("填写说明")
    wb.save(path)
    return path


def test_import_normal_question_types(tmp_path):
    path = make_xlsx(tmp_path / "ok.xlsx", [
        ["单选题", "HTML 段落标签？", "<p>", "<br>", "", "", "", "", "A", "", "HTML", "简单"],
        ["多选题", "哪些是标签？", "<p>", "<div>", "color", "<h1>", "", "", "A,B,D", "", "HTML", "中等"],
        ["判断题", "CSS 可以设置样式。", "", "", "", "", "", "", "对", "", "CSS", "简单"],
        ["填空题", "HTML 扩展名？", "", "", "", "", "", "", "html|htm", "", "HTML", "简单"],
    ])
    result = ExcelImportService().import_file(path)
    assert result.total_count == 4
    assert result.valid_count == 4
    assert result.questions[1].correct_answer == "ABD"
    assert result.questions[2].correct_answer == "正确"


def test_import_validation_errors(tmp_path):
    path = make_xlsx(tmp_path / "bad.xlsx", [
        ["单选题", "缺少答案", "A", "B", "", "", "", "", "", "", "", ""],
        ["单选题", "选项不足", "A", "", "", "", "", "", "A", "", "", ""],
        ["单选题", "错误答案字母", "A", "B", "", "", "", "", "C", "", "", ""],
        ["判断题", "判断题答案错误", "", "", "", "", "", "", "也许", "", "", ""],
        ["填空题", "重复题", "", "", "", "", "", "", "x", "", "", ""],
        ["填空题", "重复题", "", "", "", "", "", "", "x", "", "", ""],
    ])
    result = ExcelImportService().import_file(path)
    assert result.error_count == 5
    assert result.duplicate_count == 1
    text = "\n".join(result.errors)
    assert "正确答案为空" in text
    assert "至少需要两个选项" in text
    assert "必须对应已有选项" in text
    assert "判断题答案无法识别" in text
    assert "存在重复题目" in text


def test_missing_sheet(tmp_path):
    wb = Workbook()
    wb.active.title = "错误名称"
    path = tmp_path / "missing.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="题库数据"):
        ExcelImportService().import_file(path)
