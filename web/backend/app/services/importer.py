from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from tempfile import NamedTemporaryFile

from fastapi import UploadFile
from openpyxl import load_workbook

from app.services.answer_utils import normalize_answer, normalize_judge_answer, normalize_question_type

HEADERS = ["题型", "题目", "A选项", "B选项", "C选项", "D选项", "E选项", "F选项", "正确答案", "解析", "章节", "难度"]
OPTION_KEYS = ("A", "B", "C", "D", "E", "F")


@dataclass
class ImportedQuestion:
    row_number: int
    question_type: str
    question_text: str
    options: dict[str, str]
    correct_answer: str
    analysis: str = ""
    chapter: str = ""
    difficulty: str = ""
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.errors


def validate_question(q: ImportedQuestion, seen_texts: set[str] | None = None) -> ImportedQuestion:
    q.errors.clear()
    q.question_type = normalize_question_type(q.question_type)
    q.question_text = (q.question_text or "").strip()
    q.correct_answer = normalize_answer(q.question_type, q.correct_answer)
    q.options = {k: (v or "").strip() for k, v in q.options.items() if (v or "").strip()}
    q.analysis = (q.analysis or "").strip()
    q.chapter = (q.chapter or "").strip()
    q.difficulty = (q.difficulty or "").strip()
    if not q.question_text:
        q.errors.append("题目内容为空")
    if seen_texts is not None and q.question_text:
        if q.question_text in seen_texts:
            q.errors.append("存在重复题目")
        else:
            seen_texts.add(q.question_text)
    if q.question_type not in {"单选题", "多选题", "判断题", "填空题"}:
        q.errors.append("题型不支持")
    if not q.correct_answer:
        q.errors.append("正确答案为空")
    if q.question_type == "单选题":
        if len(q.options) < 2:
            q.errors.append("单选题至少需要两个选项")
        if len(q.correct_answer) != 1 or not q.correct_answer.isalpha():
            q.errors.append("单选题正确答案不能填写多个选项")
        elif q.correct_answer not in q.options:
            q.errors.append("单选题正确答案必须对应已有选项")
    elif q.question_type == "多选题":
        if len(q.options) < 2:
            q.errors.append("多选题至少需要两个选项")
        if len(q.correct_answer) < 2:
            q.errors.append("多选题正确答案至少包含两个选项")
        for letter in q.correct_answer:
            if letter not in q.options:
                q.errors.append(f"多选题答案 {letter} 没有对应选项")
                break
    elif q.question_type == "判断题":
        normalized = normalize_judge_answer(q.correct_answer)
        if normalized is None:
            q.errors.append("判断题答案无法识别")
        else:
            q.correct_answer = normalized
    return q


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


async def import_excel_upload(file: UploadFile) -> dict:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise ValueError("只支持 .xlsx 格式的 Excel 文件")
    content = await file.read()
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(content)
        tmp_path = Path(tmp.name)
    try:
        wb = load_workbook(tmp_path, data_only=True, read_only=True)
    finally:
        tmp_path.unlink(missing_ok=True)
    if "题库数据" not in wb.sheetnames:
        raise ValueError("缺少工作表“题库数据”，请使用标准模板")
    ws = wb["题库数据"]
    header = [_cell_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = [item for item in HEADERS if item not in header]
    if missing:
        raise ValueError(f"表头缺失：{', '.join(missing)}")
    indexes = {name: header.index(name) for name in HEADERS}
    questions: list[ImportedQuestion] = []
    seen: set[str] = set()
    errors: list[str] = []
    duplicates = 0
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(_cell_text(v) == "" for v in row):
            continue
        def value(name: str) -> str:
            idx = indexes[name]
            return _cell_text(row[idx] if idx < len(row) else "")
        q = ImportedQuestion(
            row_number=row_number,
            question_type=value("题型"),
            question_text=value("题目"),
            options={key: value(f"{key}选项") for key in OPTION_KEYS},
            correct_answer=value("正确答案"),
            analysis=value("解析"),
            chapter=value("章节"),
            difficulty=value("难度"),
        )
        validate_question(q, seen)
        if any("重复题目" in err for err in q.errors):
            duplicates += 1
        errors.extend(f"第 {row_number} 行：{err}" for err in q.errors)
        questions.append(q)
    if not questions:
        raise ValueError("Excel 中没有可读取的题目")
    return {
        "file_name": file.filename,
        "total_count": len(questions),
        "valid_count": sum(q.is_valid for q in questions),
        "error_count": sum(not q.is_valid for q in questions),
        "duplicate_count": duplicates,
        "errors": errors,
        "questions": [q.__dict__ | {"is_valid": q.is_valid} for q in questions],
    }
